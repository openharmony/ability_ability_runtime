#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#  Copyright (c) 2026 Huawei Device Co., Ltd.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
"""
Generic deep-scan Excel report generator.

Reads a JSON file containing consolidated findings from the three deep-scan
layers (high-impact-bug-audit, logic-analyzer, security-review) and produces
a styled .xlsx workbook that strictly follows the template defined in
.claude/skills/deep-scan/SKILL.md.

Usage:
    python3 generate_report.py \
        --module-name <module_name> \
        --input <module_name>_deep_scan_findings.json \
        [--output-dir <dir>]

Input JSON schema:
    {
      "findings": [
        {
          "file_path": "relative/path/file.cpp",
          "line_number": "123 或 120-130",
          "summary": "问题概述",
          "description": "详细描述（支持 Markdown）",
          "issue_type": "问题类型",
          "risk_level": "致命|严重|一般|提示"
        }
      ]
    }

Each finding may use either Chinese or English field names:
  Chinese: 文件路径, 行号, 问题概述, 问题详细描述, 问题类型, 风险等级
  English: file_path, line_number, summary, description, issue_type, risk_level

Risk levels may be either Chinese (致命/严重/一般/提示) or English
(critical/high/medium/low).  English levels are normalized to Chinese
in the generated Excel.
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------- Template constants (must match deep-scan/SKILL.md) ----------
HEADERS = ["文件路径", "行号", "问题概述", "问题详细描述", "问题类型", "风险等级"]
COLUMN_WIDTHS = {1: 55, 2: 18, 3: 40, 4: 90, 5: 18, 6: 12}

# Chinese risk levels are canonical in the output Excel.
RISK_RANK = {"致命": 0, "严重": 1, "一般": 2, "提示": 3}
VALID_RISK_LEVELS = set(RISK_RANK.keys())

# Accept English aliases in input; they are normalized to Chinese for output.
RISK_LEVEL_ALIASES = {
    "critical": "致命",
    "high": "严重",
    "medium": "一般",
    "low": "提示",
}

# Accept either Chinese or English field names in input JSON.
FIELD_ALIASES = {
    "文件路径": "file_path",
    "行号": "line_number",
    "问题概述": "summary",
    "问题详细描述": "description",
    "问题类型": "issue_type",
    "风险等级": "risk_level",
}

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)

RISK_FILL = {
    "致命": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "严重":   PatternFill(start_color="FFFF7F7F", end_color="FFFF7F7F", fill_type="solid"),
    "一般":   PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
    "提示":   PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
}

RISK_FONT = {
    "致命": Font(name="Calibri", size=11, bold=True, color="FFFFFFFF"),
    "严重":   Font(name="Calibri", size=11, bold=True),
    "一般":   Font(name="Calibri", size=11),
    "提示":   Font(name="Calibri", size=11),
}

DATA_ROW_HEIGHT = 160
HEADER_ROW_HEIGHT = 28


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a deep-scan Excel report from findings JSON."
    )
    parser.add_argument(
        "--module-name", "-m", required=True,
        help="Module name used for the worksheet title and output filename."
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="Path to the findings JSON file."
    )
    parser.add_argument(
        "--output-dir", "-o", default=".",
        help="Directory where the .xlsx file will be written (default: current directory)."
    )
    return parser.parse_args()


def _normalize_finding_keys(finding: dict) -> dict:
    """Map Chinese field names and English aliases to canonical English keys."""
    normalized: dict = {}
    for key, value in finding.items():
        canonical = FIELD_ALIASES.get(key, key)
        normalized[canonical] = value
    return normalized


def _normalize_risk_level(level: str) -> str:
    """Convert English risk levels to canonical Chinese levels."""
    if not isinstance(level, str):
        return level
    lower = level.strip().lower()
    return RISK_LEVEL_ALIASES.get(lower, level.strip())


def load_findings(path: str) -> list[dict]:
    p = Path(path)
    if not p.exists():
        print(f"Error: Input file not found: {path}", file=sys.stderr)
        sys.exit(1)

    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as exc:
        print(f"Error: Invalid JSON in {path}: {exc}", file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, dict):
        print("Error: JSON root must be an object", file=sys.stderr)
        sys.exit(1)
    if "findings" not in data:
        print("Error: JSON root must contain a 'findings' array", file=sys.stderr)
        sys.exit(1)
    findings = data["findings"]
    if not isinstance(findings, list):
        print("Error: 'findings' must be an array", file=sys.stderr)
        sys.exit(1)
    return findings


def validate_finding(finding: dict, index: int) -> dict | None:
    required = {"file_path", "line_number", "summary", "description", "issue_type", "risk_level"}
    missing = required - finding.keys()
    if missing:
        print(
            f"Warning: Finding {index} missing fields {sorted(missing)}, skipping",
            file=sys.stderr,
        )
        return None

    if not isinstance(finding.get("file_path"), str) or not finding["file_path"].strip():
        print(f"Warning: Finding {index} has empty or invalid file_path, skipping", file=sys.stderr)
        return None
    if not isinstance(finding.get("summary"), str) or not finding["summary"].strip():
        print(f"Warning: Finding {index} has empty or invalid summary, skipping", file=sys.stderr)
        return None

    level = finding.get("risk_level")
    normalized_level = _normalize_risk_level(level)
    if normalized_level not in VALID_RISK_LEVELS:
        print(
            f"Warning: Finding {index} has invalid risk_level '{level}', skipping",
            file=sys.stderr,
        )
        return None

    finding["risk_level"] = normalized_level
    return finding


def sort_findings(findings: list[dict]) -> list[dict]:
    def sort_key(finding: dict):
        return (
            RISK_RANK.get(finding["risk_level"], 99),
            finding.get("file_path", ""),
            str(finding.get("line_number", "")),
        )
    return sorted(findings, key=sort_key)


def build_workbook(module_name: str, findings: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{module_name} 高影响问题"

    # Column widths
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Data rows
    for row_idx, finding in enumerate(findings, start=2):
        values = [
            finding["file_path"],
            finding["line_number"],
            finding["summary"],
            finding["description"],
            finding["issue_type"],
            finding["risk_level"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 6:  # 风险等级 column
                level = finding["risk_level"]
                cell.font = RISK_FONT.get(level, DATA_FONT)
                cell.fill = RISK_FILL.get(level, PatternFill())
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

    ws.freeze_panes = "A2"
    wb.save(str(output_path))
    print(f"Saved: {output_path}, total issues: {len(findings)}")


def main() -> None:
    args = parse_args()
    module_name = args.module_name.strip()
    if not module_name:
        print("Error: --module-name must be a non-empty string", file=sys.stderr)
        sys.exit(1)

    raw_findings = load_findings(args.input)
    normalized_findings = [_normalize_finding_keys(rf) for rf in raw_findings]
    valid_findings = [
        f for f in (validate_finding(rf, i) for i, rf in enumerate(normalized_findings, 1))
        if f is not None
    ]
    sorted_findings = sort_findings(valid_findings)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{module_name}_deep_scan_issues.xlsx"

    build_workbook(module_name, sorted_findings, output_path)


if __name__ == "__main__":
    main()
