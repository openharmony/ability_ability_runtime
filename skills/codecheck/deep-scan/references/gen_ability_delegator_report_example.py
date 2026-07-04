#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#  Copyright (c) 2026 Huawei Device Co., Ltd.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
"""
Historical example: hardcoded deep-scan Excel report for
frameworks/js/napi/app/ability_delegator.

For new scans, use the generic generator instead:
  .claude/skills/deep-scan/scripts/generate_report.py

Consolidates findings from three layers:
  Layer 1: High-Impact Bug Audit
  Layer 2: Logic Analyzer
  Layer 3: Security Review
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "/root/ability_ability_runtime/ability_delegator_deep_scan_issues.xlsx"

# Each row: (文件路径, 行号, 问题概述, 问题详细描述, 问题类型, 风险等级)
issues = [
    # ============== P0 / 致命 ==============
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "417-425, 448-456",
        "全局 monitor 表在 erase 后继续遍历/解引用，迭代器失效风险",
        "### 问题描述\n"
        "`OnRemoveAbilityMonitor`/`OnRemoveAbilityMonitorSync` 在 for 循环中直接执行 "
        "`g_monitorRecord.erase(iter)` 后使用 `break`，但代码风格上 `g_monitorRecord` 与 "
        "`g_interopMonitorRecord`、`g_stageMonitorRecord` 在多处循环 erase 时未对容器加互斥锁，"
        "而 `ParseMonitorPara`、`ParseStageMonitorPara`、`ParseInteropMonitorPara` 在其他线程 "
        "可能并发写同一容器。\n"
        "`g_monitorRecord` 与 `g_interopMonitorRecord` 全程没有加锁保护，并发 Add/Remove/Parse 时 "
        "会触发 std::map 迭代器失效、节点被并发释放，造成 UAF 或崩溃。\n\n"
        "### 代码片段\n"
        "```cpp\n"
        "for (auto iter = g_monitorRecord.begin(); iter != g_monitorRecord.end(); ++iter) {\n"
        "    ...\n"
        "    if (isEquals) { g_monitorRecord.erase(iter); break; }\n"
        "}\n"
        "// g_monitorRecord、g_interopMonitorRecord 全程未加锁\n"
        "```\n\n"
        "### 修复建议\n"
        "1. 与 `g_mtxStageMonitorRecord`、`g_mutexAbilityRecord` 一致，为 "
        "`g_monitorRecord`、`g_interopMonitorRecord` 各自增加 std::mutex，所有读写路径加锁；\n"
        "2. erase 后立即 break 或使用 `iter = erase(iter)` 形式，避免使用失效迭代器。\n\n"
        "### 影响\n"
        "测试框架在 Add/Remove/Parse Monitor 并发调用时可能崩溃，影响所有 "
        "AbilityDelegator 使用方的测试稳定性。",
        "并发安全",
        "致命",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1122-1142",
        "ParseAbilityPara 在加锁状态下执行 napi_strict_equals 等 JS 引擎调用，潜在死锁/重入",
        "### 问题描述\n"
        "`ParseAbilityPara` 内部对 `g_mutexAbilityRecord` 加 `unique_lock` 后，循环调用 "
        "`iter->first.lock()->GetNapiValue()` 与 `napi_strict_equals`。这些调用会进入 JS 引擎，"
        "若另一线程在 JS 回调中再调用 `AddAbilityMonitor`（间接获取同一把锁或触发 GC/回调），"
        "存在锁-JS 重入风险。\n"
        "此外循环内对每个非匹配项都做一次 `lock()->GetNapiValue()`，若 `weak_ptr` 在 erase 之后"
        "正好被并发析构（lock() 返回 nullptr），代码未判空即调用 GetNapiValue，触发空指针解引用。\n\n"
        "### 代码片段\n"
        "```cpp\n"
        "std::unique_lock<std::mutex> lck(g_mutexAbilityRecord);\n"
        "for (auto iter = g_abilityRecord.begin(); ...) {\n"
        "    if (iter->first.expired()) { iter = g_abilityRecord.erase(iter); continue; }\n"
        "    bool isEquals = false;\n"
        "    napi_strict_equals(env, value, iter->first.lock()->GetNapiValue(), &isEquals);\n"
        "    ...\n"
        "}\n"
        "```\n\n"
        "### 修复建议\n"
        "1. 先在锁内收集候选 `shared_ptr<NativeReference>` 到本地变量，再释放锁后执行 JS 调用；\n"
        "2. `iter->first.lock()` 后判空再调用 GetNapiValue。\n\n"
        "### 影响\n"
        "高并发测试场景下死锁或空指针崩溃。",
        "并发安全",
        "致命",
    ),

    # ============== P1 / 高 ==============
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1061-1065, 1161-1164, 1539-1546, 1670-1674",
        "napi_create_reference 返回值未校验，失败时 ref 为 nullptr 被强转为 NativeReference* 入表",
        "### 问题描述\n"
        "多处模式：\n"
        "```cpp\n"
        "napi_ref ref = nullptr;\n"
        "napi_create_reference(env, value, 1, &ref);   // 未检查 napi_ok 与 ref\n"
        "reference.reset(reinterpret_cast<NativeReference*>(ref));\n"
        "g_monitorRecord.emplace(reference, monitor);\n"
        "```\n"
        "若 `napi_create_reference` 失败（OOM、value 非对象等），`ref` 仍为 nullptr，"
        "`reinterpret_cast<NativeReference*>(nullptr)` 得到的 shared_ptr 指向裸 nullptr，"
        "后续 `jsMonitor->GetNapiValue()` 解引用 nullptr 崩溃，或污染全局表（以 nullptr 为 key）。\n\n"
        "### 修复建议\n"
        "1. 检查 `napi_create_reference` 返回值是否为 `napi_ok` 且 `ref != nullptr`；\n"
        "2. 失败时返回 nullptr 并向 JS 抛异常，不要将空 reference 入表。\n\n"
        "### 影响\n"
        "异常路径下后续 Add/Remove/Wait Monitor 调用解引用空指针，进程崩溃。",
        "错误处理",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1028-1038, 1089-1110, 1642-1652",
        "napi_get_named_property 返回值未校验，ConvertFromJsValue 失败路径不统一",
        "### 问题描述\n"
        "`ParseMonitorPara`/`ParseStageMonitorPara`/`ParseInteropMonitorPara` 调用 "
        "`napi_get_named_property(env, value, \"abilityName\", &abilityNameValue)` 后直接判 "
        "`abilityNameValue == nullptr`。但 NAPI 规范下 `napi_get_named_property` 即使属性不存在通常"
        "也返回 napi_ok 且写入 undefined（非 nullptr），仅当 env/value 非法时才可能写入 nullptr。"
        "因此当前判空无法拦截 “属性缺失” 场景，进而 `ConvertFromJsValue(undefined)` 返回 false，"
        "函数返回 nullptr 被上游当作参数错误处理。\n"
        "更关键的是 `moduleName` 路径：当 `napi_get_named_property` 写入 undefined 时，"
        "`moduleNameValue != nullptr` 成立，进入 ConvertFromJsValue 失败分支被静默忽略 "
        "（`moduleName = \"\"`），与文档要求不符但不致命；然而 `abilityName` 路径未做 "
        "`napi_typeof` 校验，攻击者可传入 `{ abilityName: 123 }` 等非字符串导致 "
        "ConvertFromJsValue 失败。\n\n"
        "### 修复建议\n"
        "1. 使用 `napi_typeof` / `CheckTypeForNapiValue` 显式校验 abilityName 为 napi_string；\n"
        "2. 校验 `napi_get_named_property` 返回值为 `napi_ok`；\n"
        "3. 区分 “属性缺失” 与 “类型不符” 的错误码。\n\n"
        "### 影响\n"
        "参数校验不严格，可能导致 Monitor 注册到空 abilityName，"
        "后续无法匹配实际 Ability，测试逻辑失效。",
        "输入验证",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "51-56",
        "全局 monitor/ability 记录表生命周期长于 JS 对象，存在内存泄漏与 stale 引用",
        "### 问题描述\n"
        "`g_monitorRecord`、`g_stageMonitorRecord`、`g_interopMonitorRecord`、`g_abilityRecord` 均为 "
        "进程级全局 std::map，key 为 `shared_ptr<NativeReference>` 或 `weak_ptr<NativeReference>`。"
        "其中 `g_monitorRecord`、`g_interopMonitorRecord` 使用 shared_ptr 持有 NativeReference，"
        "这意味着即使 JS 侧 Monitor 对象已被 GC 回收（没有显式调用 removeAbilityMonitor），"
        "C++ 侧仍强引用其 NativeReference，导致 JS 对象永远无法释放 → 内存泄漏。\n"
        "另一方面，`g_abilityRecord` 使用 weak_ptr 是合理的，但 `g_monitorRecord`/`g_interopMonitorRecord` "
        "对 shared_ptr 的使用没有配套的 GC 清理逻辑（仅在 Remove 接口中按值匹配删除），"
        "若测试用例未成对调用 Add/Remove，monitor 表会无限增长。\n\n"
        "### 修复建议\n"
        "1. 将 `g_monitorRecord`、`g_interopMonitorRecord` 的 key 改为 weak_ptr 并配套自定义 owner_less；\n"
        "2. 或在 JS 对象 Finalizer 中主动清理对应记录；\n"
        "3. 定期遍历清理 expired 项。\n\n"
        "### 影响\n"
        "长时间运行的测试进程内存持续增长，最终 OOM；跨测试用例的状态污染。",
        "资源泄漏",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "693-749",
        "ExecuteShellCommand 将用户传入字符串原样转交底层执行，缺乏命令注入缓解",
        "### 问题描述\n"
        "`OnExecuteShellCommand` 接收 JS 传入的 `cmd` 字符串，仅做 ConvertFromJsValue 读取后即"
        "通过 `delegator->ExecuteShellCommand(cmd, timeoutSecs)` 转发执行。当前代码层未做：\n"
        "- 命令长度上限校验（可触发超大字符串处理）；\n"
        "- timeout 合法范围校验（负数、极大值导致永久阻塞）；\n"
        "- cmd 是否包含危险字符的告警/审计。\n"
        "虽然 AbilityDelegator 通常运行在测试进程、命令最终由受控的 aa 工具执行，但在跨应用测试 "
        "场景下，被测应用若能调用此接口，可借助测试框架的特权执行任意 shell。\n\n"
        "### 代码片段\n"
        "```cpp\n"
        "if (!ConvertFromJsValue(env, info.argv[INDEX_ZERO], cmd)) { return nullptr; }\n"
        "// 未校验 cmd.size()、cmd 内容、timeout 范围\n"
        "shellCmdResultBox->shellCmdResult_ = delegator->ExecuteShellCommand(cmd, timeoutSecs);\n"
        "```\n\n"
        "### 修复建议\n"
        "1. 在 NAPI 层限制 cmd 最大长度（如 4096）；\n"
        "2. 校验 timeoutSecs 范围（>= 0 且 <= 上限）；\n"
        "3. 对 cmd 内容做基础白名单/转义审计；\n"
        "4. 在 HiLog 中记录调用者 bundle 与 cmd 摘要以便审计。\n\n"
        "### 影响\n"
        "若 AbilityDelegator 被非预期调用方触达，存在命令注入/特权执行风险。",
        "权限安全",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_interop_ability_monitor.cpp",
        "113-126",
        "std::static_pointer_cast<EtsDelegatorAbilityProperty> 未做运行时类型校验",
        "### 问题描述\n"
        "`ConvertAbilityToNapiValue` 中：\n"
        "```cpp\n"
        "auto property = abilityObj.lock();\n"
        "if (property == nullptr) return CreateJsNull(env_);\n"
        "auto etsProperty = std::static_pointer_cast<AppExecFwk::EtsDelegatorAbilityProperty>(property);\n"
        "if (etsProperty == nullptr) { ... }\n"
        "```\n"
        "`std::static_pointer_cast` 不会执行运行时类型检查，若底层对象实际为 "
        "`ADelegatorAbilityProperty`（非 ETS 路径），指针仍非空，但访问 `etsProperty->object_` "
        "（期望 ETS 专用 ani_ref 字段）会读到错误偏移，引发未定义行为。\n"
        "`if (etsProperty == nullptr)` 永远为 false（static_pointer_cast 不返回空），"
        "该判空是死代码，给人虚假的安全感。\n\n"
        "### 修复建议\n"
        "1. 改用 `std::dynamic_pointer_cast<EtsDelegatorAbilityProperty>(property)`；\n"
        "2. 失败时回退 CreateJsNull 并打 ERROR 日志；\n"
        "3. 确保 BaseDelegatorAbilityProperty 具备虚析构和 RTTI（至少含一个虚函数）。\n\n"
        "### 影响\n"
        "非 ETS Ability 误入 interop 回调路径时，读取非法内存偏移导致崩溃或数据泄漏。",
        "类型安全",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/ability_monitor.cpp",
        "44, 63, 82, 101, 121, 140, 159",
        "std::static_pointer_cast<ADelegatorAbilityProperty> 无运行时类型校验",
        "### 问题描述\n"
        "`AbilityMonitor::OnAbilityStart` 等回调中，对 `baseProperty` 仅判空后即 "
        "`std::static_pointer_cast<ADelegatorAbilityProperty>(baseProperty)`。与上一条同理，"
        "若实际派生类型不符（例如误被 ETS 路径复用该 monitor），static_cast 不做校验，"
        "后续 `jsbaseProperty->object_`（weak_ptr<NativeReference>）访问的是错误内存偏移。\n"
        "此外 `OnAbilityStart` 在 lock 成功后才检查 `jsMonitor_ == nullptr`，"
        "顺序上若 jsMonitor_ 为空，前面 static_cast 已产生悬空指针引用风险。\n\n"
        "### 修复建议\n"
        "1. 改为 dynamic_pointer_cast 并校验；\n"
        "2. 先校验 jsMonitor_ 再做 cast，减少无用工作。\n\n"
        "### 影响\n"
        "跨语言/跨模型场景下类型不符导致 UB。",
        "类型安全",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "99-127, 752-785",
        "AttachAppContext / OnGetAppContext 中 napi_wrap 失败后 double-free workContext 风险",
        "### 问题描述\n"
        "`AttachAppContext` 与 `OnGetAppContext` 中：\n"
        "```cpp\n"
        "auto workContext = new (std::nothrow) std::weak_ptr<...>(ptr);\n"
        "napi_coerce_to_native_binding_object(env, object, DetachCallbackFunc, AttachAppContext, value, nullptr);\n"
        "napi_status status = napi_wrap(env, object, workContext, finalizer, nullptr, nullptr);\n"
        "if (status != napi_ok && workContext != nullptr) {\n"
        "    delete workContext;\n"
        "    return nullptr;\n"
        "}\n"
        "```\n"
        "`napi_coerce_to_native_binding_object` 已经把 workContext 通过 hint 传入；若此后 "
        "`napi_wrap` 失败，代码 `delete workContext` 后返回 nullptr，但 "
        "`napi_coerce_to_native_binding_object` 内部可能已注册 detach 回调并引用该指针；"
        "JS 引擎后续触发 detach 时会使用已释放指针，存在 UAF 风险。\n"
        "另外 `AttachAppContext` 中 `value`（void*）来自上层 weak_ptr 指针，"
        "`napi_coerce_to_native_binding_object` 第 5 参数传 workContext，而 finalizer 在 wrap 失败后"
        "不会被调用，由谁来释放 workContext 取决于引擎实现，存在不确定性。\n\n"
        "### 修复建议\n"
        "1. 在调用 napi_wrap 前不要先调用 napi_coerce_to_native_binding_object；\n"
        "2. 失败路径统一在 finalizer 之外手动管理；\n"
        "3. 使用 unique_ptr 在栈上托管，napi_wrap 成功后 release()。\n\n"
        "### 影响\n"
        "GC/detach 时刻触发 UAF，导致进程崩溃或内存损坏。",
        "内存安全",
        "高",
    ),

    # ============== P1-2 / 高-中 ==============
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "129-153",
        "JSAbilityDelegator 构造函数捕获 clearFunc 引用全局 g_abilityRecord，delegator 生命周期不可控",
        "### 问题描述\n"
        "构造函数在拿到 delegator 后立即 `delegator->RegisterClearFunc(clearFunc)`，"
        "clearFunc 内部访问全局 `g_abilityRecord` 和 `g_mutexAbilityRecord`。"
        "由于 JSAbilityDelegator 是每个 napi 对象独立构造（可能多个），"
        "RegisterClearFunc 会被多次调用，底层若未做去重，clearFunc 会重复注册；"
        "而 clearFunc 又是引用全局静态变量，多次注册 + 并发 clear 时锁竞争与迭代器失效叠加。\n"
        "另外 `std::static_pointer_cast<ADelegatorAbilityProperty>(baseProperty)` 同样存在类型校验缺失。\n\n"
        "### 修复建议\n"
        "1. RegisterClearFunc 应支持去重或只注册一次（用 once_flag）；\n"
        "2. clearFunc 内部对 cast 做判空。\n\n"
        "### 影响\n"
        "重复 clear 注册导致同一 ability 被多次清理；并发 erase 触发迭代器失效。",
        "逻辑缺陷",
        "高",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "537-562, 598-619",
        "WaitAbilityMonitor execute callback 中未捕获 delegator 调用异常，失败路径状态不一致",
        "### 问题描述\n"
        "`OnWaitAbilityMonitor` 的 execute lambda：\n"
        "```cpp\n"
        "std::shared_ptr<BaseDelegatorAbilityProperty> property = opt.hasTimeoutPara ?\n"
        "    delegator->WaitAbilityMonitor(monitor, timeout) : delegator->WaitAbilityMonitor(monitor);\n"
        "auto jsProperty = std::static_pointer_cast<ADelegatorAbilityProperty>(property);\n"
        "if (!jsProperty || jsProperty->object_.expired()) { return; }\n"
        "abilityObjectBox->object_ = jsProperty->object_;\n"
        "{\n"
        "    std::unique_lock<std::mutex> lck(g_mutexAbilityRecord);\n"
        "    g_abilityRecord.emplace(jsProperty->object_, jsProperty->token_);\n"
        "}\n"
        "```\n"
        "若 `WaitAbilityMonitor` 内部抛 C++ 异常（如 IPC 错误），execute lambda 未捕获，"
        "异常会跨越 NAPI 边界，行为未定义（通常 crash）。"
        "另外 object_ 是 weak_ptr，emplace 后若 JS 侧立即释放，map 中留下 stale weak 项，"
        "下次 ParseAbilityPara 才会清理，期间 token_（sptr）依然被持有，导致 ability 无法真正释放。\n\n"
        "### 修复建议\n"
        "1. 在 execute lambda 外层包 try/catch；\n"
        "2. emplace 后立即检查 weak 是否 expired，若是则不写入。\n\n"
        "### 影响\n"
        "IPC 失败时进程崩溃；测试期间 ability 句柄泄漏。",
        "错误处理",
        "高",
    ),

    # ============== P2 / 中 ==============
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1467-1471",
        "ParseMockListPara 中 napi_get_property_names 返回值未校验",
        "### 问题描述\n"
        "```cpp\n"
        "napi_value array = nullptr;\n"
        "napi_get_property_names(env, value, &array);\n"
        "if (!ParseArrayStringValue(env, array, propNames)) { return false; }\n"
        "```\n"
        "`napi_get_property_names` 返回 napi_status 未校验；若返回非 napi_ok 且 array 仍为 nullptr，"
        "虽然 ParseArrayStringValue 内部判空，但失败原因被掩盖为 'invalid propNames'，"
        "与真正错误（引擎错误）无法区分。此外 mockList 直接被写入 NativeEngine->SetMockModuleList，"
        "若 propName 中包含特殊字符（如 '__proto__'）可能引发原型链污染语义。\n\n"
        "### 修复建议\n"
        "1. 校验 napi_get_property_names 返回 napi_ok；\n"
        "2. 过滤危险 key（含 __proto__/constructor 等）。\n\n"
        "### 影响\n"
        "Mock 配置异常导致测试模块路径错误。",
        "输入验证",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1496-1526",
        "ParseArrayStringValue 在 ConvertFromJsValue 失败时 continue 而非 fail，行为不一致",
        "### 问题描述\n"
        "```cpp\n"
        "for (uint32_t i = 0; i < arrayLen; i++) {\n"
        "    napi_get_element(env, array, i, &jsValue);\n"
        "    if (!ConvertFromJsValue(env, jsValue, strItem)) {\n"
        "        TAG_LOGW(...); continue;\n"
        "    }\n"
        "    vector.emplace_back(std::move(strItem));\n"
        "}\n"
        "return true;\n"
        "```\n"
        "数组中混入非字符串元素时静默跳过，调用方拿到截断后的列表却认为成功，"
        "导致 mockList 部分丢失但不报错。与 `ParseStageMonitorPara` 中 ConvertFromJsValue 失败即返回 "
        "nullptr 的策略不一致。\n\n"
        "### 修复建议\n"
        "1. 统一策略：要么全部严格校验、要么全部宽松；\n"
        "2. 至少在 Debug 构建中统计跳过数量并 ERROR。\n\n"
        "### 影响\n"
        "Mock 配置部分丢失，测试结果不可靠。",
        "逻辑缺陷",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "875-883",
        "StartAbility 错误码处理依赖 GetApiTargetVersion 取模，版本判断边界易错",
        "### 问题描述\n"
        "```cpp\n"
        "uint32_t apiTargetVersion = delegator->GetApiTargetVersion() % API_VERSION_MOD;  // % 100\n"
        "if (apiTargetVersion >= API20) {\n"
        "    task.Reject(env, CreateJsError(env, static_cast<int>(GetJsErrorCodeByNativeError(result)), ...));\n"
        "} else {\n"
        "    task.Reject(env, CreateJsError(env, result, ...));\n"
        "}\n"
        "```\n"
        "`API_VERSION_MOD = 100` 的取模假设版本号格式为 `major*100 + minor`，"
        "若未来 API 版本编号方案变更（如直接用 12, 13, 14...），取模逻辑会错误地 "
        "把高版本判为低版本，返回旧风格错误码，破坏向前兼容。\n\n"
        "### 修复建议\n"
        "1. 不取模，直接比较 GetApiTargetVersion 与阈值；\n"
        "2. 或将版本号语义在公共头文件中显式文档化。\n\n"
        "### 影响\n"
        "未来版本号方案调整后错误码回归，影响应用兼容性。",
        "逻辑缺陷",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1293-1327",
        "ParseTimeoutCallbackPara 处理 argv[1]/argv[2] 时对 undefined/number 边界处理含糊",
        "### 问题描述\n"
        "```cpp\n"
        "if (info.argc >= ARGC_TWO) {\n"
        "    if (ConvertFromJsValue(env, info.argv[INDEX_ONE], timeout)) {\n"
        "        opt.hasTimeoutPara = true;\n"
        "    } else {\n"
        "        if (info.argv[INDEX_ONE] == nullptr) { TAG_LOGW(...); }\n"
        "        else if (IsTypeForNapiValue(... napi_function)) { opt.hasCallbackPara = true; return ...; }\n"
        "        else { return nullptr; }\n"
        "    }\n"
        "    if (info.argc > ARGC_TWO) {\n"
        "        if (!IsTypeForNapiValue(env, info.argv[INDEX_TWO], napi_function)) {\n"
        "            if (info.argv[INDEX_TWO] == nullptr) { return CreateJsNull(env); }\n"
        "            return nullptr;\n"
        "        }\n"
        "        opt.hasCallbackPara = true;\n"
        "    }\n"
        "}\n"
        "```\n"
        "当 `argv[1]` 是 undefined（不是 nullptr）时，ConvertFromJsValue 失败，进入 else 分支，"
        "既不是 nullptr 也不是 function，会 `return nullptr`（参数错误）。"
        "而 JS 中显式传 undefined 表示 '不传'，应被容忍。这与 TS d.ts 中 timeout 可选语义不一致。\n"
        "另外 argv[1] 是 number 但 < 0（负值）时也被接受，后续作为超时秒数传给 IPC 可能触发未定义行为。\n\n"
        "### 修复建议\n"
        "1. 显式使用 napi_typeof 区分 undefined 与 number；\n"
        "2. 校验 timeout >= 0。\n\n"
        "### 影响\n"
        "正常 JS 调用被错误判为参数错误；负超时传入底层。",
        "逻辑缺陷",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1612-1621, 1573-1596",
        "OnRemoveInteropAbilityMonitorSync 在 delegator 不存在时仍返回 undefined 不抛错",
        "### 问题描述\n"
        "```cpp\n"
        "auto delegator = AbilityDelegatorRegistry::GetAbilityDelegator(... JS);\n"
        "if (!delegator) {\n"
        "    ThrowError(env, COMMON_FAILED, \"Calling RemoveInteropAbilityMonitorSync failed.\");\n"
        "    return CreateJsUndefined(env);\n"
        "}\n"
        "for (auto iter = g_interopMonitorRecord.begin(); ...) {\n"
        "    ...\n"
        "    if (isEquals) {\n"
        "        delegator->RemoveInteropAbilityMonitor(iter->second);\n"
        "        g_interopMonitorRecord.erase(iter);\n"
        "        break;\n"
        "    }\n"
        "}\n"
        "return CreateJsUndefined(env);\n"
        "```\n"
        "若 monitor 在表中找不到（用户重复 remove 或传错对象），代码静默返回 undefined，"
        "JS 侧无法感知 'remove 失败'，与 Add 接口的 ThrowError 行为不一致。\n"
        "此外循环未加锁（参见前述并发问题）。\n\n"
        "### 修复建议\n"
        "1. 未找到时通过 ThrowError 或返回 false 让 JS 感知；\n"
        "2. 加锁 g_interopMonitorRecord。\n\n"
        "### 影响\n"
        "测试用例逻辑错误无法及时暴露。",
        "错误处理",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_monitor.cpp",
        "74-80",
        "SetJsAbilityMonitor 中 napi_create_reference 未校验且 env_ 来自外部",
        "### 问题描述\n"
        "```cpp\n"
        "void JSAbilityMonitor::SetJsAbilityMonitor(napi_value jsAbilityMonitor) {\n"
        "    napi_ref ref = nullptr;\n"
        "    napi_create_reference(env_, jsAbilityMonitor, 1, &ref);\n"
        "    jsAbilityMonitor_ = std::unique_ptr<NativeReference>(reinterpret_cast<NativeReference*>(ref));\n"
        "}\n"
        "```\n"
        "`env_` 由 `SetJsAbilityMonitorEnv` 在外部设置，若调用顺序错误（先调 SetJsAbilityMonitor 后调 "
        "SetJsAbilityMonitorEnv），env_ 为 nullptr，napi_create_reference(nullptr, ...) 行为未定义。"
        "ref 未判空即 reinterpret_cast 入 unique_ptr 管理。\n\n"
        "### 修复建议\n"
        "1. 合并两个 Set 方法，强制 env 与 value 同时传入；\n"
        "2. 校验 env_ 非空、napi_create_reference 返回 napi_ok。\n\n"
        "### 影响\n"
        "调用顺序错误导致空 env 解引用；ref 为空导致后续 GetNapiValue 空指针解引用。",
        "错误处理",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_monitor.cpp",
        "108-115",
        "CallLifecycleCBFunction 在 abilityObj 过期时仍使用 CreateJsNull 作为回退参数",
        "### 问题描述\n"
        "```cpp\n"
        "auto nativeAbilityObj = CreateJsNull(env_);\n"
        "if (!abilityObj.expired()) {\n"
        "    nativeAbilityObj = abilityObj.lock()->GetNapiValue();\n"
        "}\n"
        "napi_value argv[] = { nativeAbilityObj };\n"
        "napi_call_function(env_, obj, method, ArraySize(argv), argv, &callResult);\n"
        "```\n"
        "`abilityObj.lock()->GetNapiValue()` 可能返回 nullptr（NativeReference 已失效但 weak_ptr 未过期），"
        "此时 argv[0] 为 nullptr，napi_call_function 收到 null argument 行为未定义（部分引擎 crash）。\n\n"
        "### 修复建议\n"
        "GetNapiValue 返回 nullptr 时回退 CreateJsNull(env_)。\n\n"
        "### 影响\n"
        "生命周期边界场景下 JS 回调收到非法参数。",
        "内存安全",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_interop_ability_monitor.cpp",
        "125-126",
        "reinterpret_cast<ani_env*>(aniEnvVoid_) 未做类型校验",
        "### 问题描述\n"
        "`aniEnvVoid_` 来自 `AbilityDelegatorRegistry::GetAniEnv()`（返回 void*），"
        "代码中 `reinterpret_cast<ani_env *>(aniEnvVoid_)` 直接转换并交给 JsInteropObject。"
        "若底层运行时不是 ETS/ArkTS，GetAniEnv 返回的指针可能并非 ani_env，"
        "后续 JsInteropObject 内部调用 ani 接口会触发 UB。\n"
        "虽然前文已判 `aniEnvVoid_ == nullptr`，但类型不匹配无法检测。\n\n"
        "### 修复建议\n"
        "1. 在 SetAniEnv 时记录 runtime 类型；\n"
        "2. 调用前校验当前 runtime 是否为 ETS/ArkTS。\n\n"
        "### 影响\n"
        "非 ETS 环境误调 interop 接口导致 UB。",
        "类型安全",
        "中",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "30, 64",
        "thread_local reference 在多 env/多 worker 场景下与全局状态不一致",
        "### 问题描述\n"
        "`js_ability_delegator_registry.cpp:30` 中 `thread_local std::unique_ptr<NativeReference> reference;`"
        "缓存了 AbilityDelegator 的 napi 引用。但 `g_monitorRecord`/`g_abilityRecord` 等是进程级全局，"
        "若同一进程存在多个 JS 线程（worker、多 worker），各线程缓存的 reference 指向的 "
        "JSAbilityDelegator napi 对象，其内部绑定的 NativeEngine 可能不同；"
        "而所有线程共享同一 g_monitorRecord，导致 A 线程注册的 monitor（env=A）"
        "在 B 线程的 GetNapiValue 调用中被使用，跨 env 操作 napi_value 属于 UB。\n\n"
        "### 修复建议\n"
        "1. 全局表按 env 隔离；\n"
        "2. 或限制 AbilityDelegator 只能在主线程使用并在文档中明确。\n\n"
        "### 影响\n"
        "多线程测试场景下跨 env 操作 napi_value 引发崩溃。",
        "并发安全",
        "中",
    ),

    # ============== P2 / 低 ==============
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "72-80",
        "ThrowJsError 在非 ENABLE_ERRCODE 构建下不抛错仅返回 undefined",
        "### 问题描述\n"
        "```cpp\n"
        "napi_value ThrowJsError(napi_env env, int32_t errCode, std::string errMsg) {\n"
        "#ifdef ENABLE_ERRCODE\n"
        "    napi_throw(env, CreateJsError(env, errCode, errMsg));\n"
        "#endif\n"
        "    return CreateJsUndefined(env);\n"
        "}\n"
        "```\n"
        "在未定义 ENABLE_ERRCODE 时，所有参数错误都被吞掉，JS 侧拿到 undefined 无法感知错误，"
        "与 d.ts 中 'throws BusinessError' 的契约不符。这是已知的兼容性策略，但属于潜在安全/可靠性问题："
        "攻击者可利用 '错误不抛出' 行为继续执行后续逻辑。\n\n"
        "### 修复建议\n"
        "至少在日志中记录被吞掉的错误码与调用栈，便于排查。\n\n"
        "### 影响\n"
        "非 ERRCODE 构建下错误被静默吞掉，难以定位。",
        "错误处理",
        "低",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator.cpp",
        "1006-1007",
        "OnSetMockList 直接 reinterpret_cast<NativeEngine*>(env) 未校验",
        "### 问题描述\n"
        "```cpp\n"
        "auto engine = reinterpret_cast<NativeEngine*>(env);\n"
        "engine->SetMockModuleList(mockList);\n"
        "```\n"
        "依赖 napi_env 与 NativeEngine* 一一对应的隐含约定。"
        "若未来 napi 实现变更（如增加间接层），reinterpret_cast 得到非法指针，"
        "SetMockModuleList 调用立即崩溃。代码缺少静态断言保证该约定。\n\n"
        "### 修复建议\n"
        "通过 NativeEngine::FromNapiEnv 之类的官方转换接口获取，避免裸 reinterpret_cast。\n\n"
        "### 影响\n"
        "引擎抽象层变更导致崩溃。",
        "代码质量",
        "低",
    ),
    (
        "frameworks/js/napi/app/ability_delegator/js_ability_delegator_registry.cpp",
        "55-72",
        "OnGetAbilityDelegator 中 napi_create_reference 未校验返回值",
        "### 问题描述\n"
        "```cpp\n"
        "napi_ref ref = nullptr;\n"
        "napi_create_reference(env, value, 1, &ref);\n"
        "reference.reset(reinterpret_cast<NativeReference*>(ref));\n"
        "```\n"
        "与其他位置同类问题：未校验 napi_create_reference 返回值与 ref 是否为空。\n\n"
        "### 修复建议\n"
        "校验返回 napi_ok 且 ref != nullptr。\n\n"
        "### 影响\n"
        "OOM 时 reference 持有空指针，后续 GetNapiValue 解引用空指针。",
        "错误处理",
        "低",
    ),
]


def build_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ability_delegator 深度扫描问题"

    # ---------- 样式 ----------
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    level_fill = {
        "致命": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
        "高":   PatternFill(start_color="FFFF7F7F", end_color="FFFF7F7F", fill_type="solid"),
        "中":   PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
        "低":   PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    }
    level_font = {
        "致命": Font(name="Calibri", size=11, bold=True, color="FFFFFFFF"),
        "高":   Font(name="Calibri", size=11, bold=True),
        "中":   Font(name="Calibri", size=11),
        "低":   Font(name="Calibri", size=11),
    }

    # 列宽
    widths = {1: 55, 2: 18, 3: 40, 4: 90, 5: 18, 6: 12}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 表头
    headers = ["文件路径", "行号", "问题概述", "问题详细描述", "问题类型", "风险等级"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 28

    # 排序：致命 > 高 > 中 > 低
    rank = {"致命": 0, "高": 1, "中": 2, "低": 3}
    sorted_issues = sorted(issues, key=lambda x: (rank.get(x[5], 9), x[0], x[1]))

    for row_idx, issue in enumerate(sorted_issues, start=2):
        path, line, summary, detail, itype, level = issue
        values = [path, line, summary, detail, itype, level]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = border
            if col_idx == 6:
                cell.font = level_font.get(level, data_font)
                cell.fill = level_fill.get(level, PatternFill())
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                cell.alignment = data_align
        ws.row_dimensions[row_idx].height = 160

    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"Saved: {path}, total issues: {len(sorted_issues)}")


if __name__ == "__main__":
    build_workbook(OUT_PATH)
